import io
from datetime import date, timedelta

from django.http import HttpResponse
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response

from core.dynamic_models import get_employee_model, get_attendance_model
from core.permissions import IsCompanyManager, BelongsToCompany


def attendance_to_dict(att):
    return {
        "id": str(att.id),
        "employee_code": att.employee_code,
        "date": str(att.date),
        "check_in": att.check_in.isoformat() if att.check_in else None,
        "check_out": att.check_out.isoformat() if att.check_out else None,
        "work_duration": str(att.work_duration) if att.work_duration else None,
        "status": att.status,
        "check_in_confidence": att.check_in_confidence,
        "check_out_confidence": att.check_out_confidence,
        "notes": att.notes,
    }


class AttendanceListView(APIView):
    permission_classes = [IsCompanyManager, BelongsToCompany]

    def get(self, request, company_code):
        Attendance = get_attendance_model(company_code.upper())
        records = Attendance.objects.all().order_by("-date", "employee_code")

        # Filters
        att_date = request.query_params.get("date")
        employee_code = request.query_params.get("employee_code")
        att_status = request.query_params.get("status")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if att_date:
            records = records.filter(date=att_date)
        if employee_code:
            records = records.filter(employee_code=employee_code)
        if att_status:
            records = records.filter(status=att_status)
        if start_date:
            records = records.filter(date__gte=start_date)
        if end_date:
            records = records.filter(date__lte=end_date)

        # Enrich with employee names
        Employee = get_employee_model(company_code.upper())
        emp_map = {e.employee_code: f"{e.first_name} {e.last_name}" for e in Employee.objects.all()}

        result = []
        for att in records:
            d = attendance_to_dict(att)
            d["employee_name"] = emp_map.get(att.employee_code, "")
            result.append(d)

        return Response({"count": len(result), "results": result})


class AttendanceSummaryView(APIView):
    """Daily summary: total present/late/absent counts."""
    permission_classes = [IsCompanyManager, BelongsToCompany]

    def get(self, request, company_code):
        summary_date = request.query_params.get("date", str(date.today()))

        Employee = get_employee_model(company_code.upper())
        Attendance = get_attendance_model(company_code.upper())

        total_employees = Employee.objects.filter(status="active").count()
        records = Attendance.objects.filter(date=summary_date)

        summary = {
            "date": summary_date,
            "total_employees": total_employees,
            "present": records.filter(status="present").count(),
            "late": records.filter(status="late").count(),
            "half_day": records.filter(status="half_day").count(),
            "absent": total_employees - records.count(),
            "total_marked": records.count(),
        }
        return Response(summary)


class AttendanceExportView(APIView):
    """Export attendance records as Excel."""
    permission_classes = [IsCompanyManager, BelongsToCompany]

    def get(self, request, company_code):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({"detail": "openpyxl not installed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        Attendance = get_attendance_model(company_code.upper())
        Employee = get_employee_model(company_code.upper())

        start_date = request.query_params.get("start_date", str(date.today() - timedelta(days=30)))
        end_date = request.query_params.get("end_date", str(date.today()))

        records = Attendance.objects.filter(
            date__gte=start_date, date__lte=end_date
        ).order_by("date", "employee_code")

        emp_map = {e.employee_code: f"{e.first_name} {e.last_name}" for e in Employee.objects.all()}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = ["Date", "Employee Code", "Employee Name", "Check In", "Check Out", "Work Duration", "Status", "Confidence"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, att in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=str(att.date))
            ws.cell(row=row_idx, column=2, value=att.employee_code)
            ws.cell(row=row_idx, column=3, value=emp_map.get(att.employee_code, ""))
            ws.cell(row=row_idx, column=4, value=att.check_in.strftime("%H:%M:%S") if att.check_in else "")
            ws.cell(row=row_idx, column=5, value=att.check_out.strftime("%H:%M:%S") if att.check_out else "")
            ws.cell(row=row_idx, column=6, value=str(att.work_duration) if att.work_duration else "")
            ws.cell(row=row_idx, column=7, value=att.status)
            ws.cell(row=row_idx, column=8, value=round(att.check_in_confidence or 0, 3))

        # Auto-size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"attendance_{company_code}_{start_date}_to_{end_date}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
